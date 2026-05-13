"""Выгрузки для админов: результаты тура в CSV / XLSX."""

from __future__ import annotations

import csv
import io
from datetime import datetime
from enum import Enum

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.db.models import (
    Round,
    RoundCode,
    User,
    UserRoundProgress,
    Winner,
    WinnerSelection,
)
from app.services.winner_ranking import fetch_ranked_participants

_SCORE_HEADERS = [
    "round_code",
    "round_status",
    "competition_rank",
    "user_id",
    "telegram_user_id",
    "tg_username",
    "email",
    "email_verified_at",
    "total_score",
    "progress_status",
    "started_at",
    "finished_at",
    "all_answers_correct",
    "winner_position",
]

_FILL_PERFECT = PatternFill(fill_type="solid", fgColor="FFFF99")


def _csv_cell(val):
    if val is None:
        return ""
    if isinstance(val, bool):
        return "1" if val else "0"
    if isinstance(val, datetime):
        return val.isoformat(sep=" ", timespec="seconds")
    if hasattr(val, "value"):
        return val.value
    return val


def _xlsx_cell(val):
    if val is None:
        return ""
    if isinstance(val, Enum):
        return val.value
    return val


async def _export_note(session: AsyncSession, rnd: Round) -> str | None:
    sel = await session.scalar(
        select(WinnerSelection).where(WinnerSelection.round_id == rnd.id)
    )
    if sel is None or not isinstance(sel.payload, dict):
        return None
    return sel.payload.get("export_note_ru")


async def _build_ordered_score_rows(session: AsyncSession, rnd: Round) -> list[tuple]:
    """Кортежи в порядке рейтинга для выгрузки."""
    ranked = await fetch_ranked_participants(session, rnd)
    if not ranked:
        return []

    ws = aliased(Winner)
    wsel = aliased(WinnerSelection)
    ids = [p.user_id for p in ranked]

    r = await session.execute(
        select(
            User.id,
            User.telegram_user_id,
            User.tg_username,
            User.email,
            User.email_verified_at,
            UserRoundProgress.total_score,
            UserRoundProgress.status,
            UserRoundProgress.started_at,
            UserRoundProgress.finished_at,
            ws.position,
        )
        .outerjoin(
            UserRoundProgress,
            (UserRoundProgress.user_id == User.id)
            & (UserRoundProgress.round_id == rnd.id),
        )
        .outerjoin(wsel, wsel.round_id == rnd.id)
        .outerjoin(
            ws,
            (ws.user_id == User.id) & (ws.winner_selection_id == wsel.id),
        )
        .where(User.id.in_(ids))
    )
    by_id = {row[0]: row for row in r.all()}

    rows_out: list[tuple] = []
    for rank, p in enumerate(ranked, start=1):
        row = by_id.get(p.user_id)
        if row is None:
            continue
        (
            _uid,
            telegram_user_id,
            tg_username,
            email,
            email_verified_at,
            total_score,
            progress_status,
            started_at,
            finished_at,
            winner_position,
        ) = row
        rows_out.append(
            (
                rank,
                rnd.code.value,
                rnd.status,
                _uid,
                telegram_user_id,
                tg_username,
                email,
                email_verified_at,
                total_score,
                progress_status,
                started_at,
                finished_at,
                p.perfect_all,
                winner_position,
            )
        )
    return rows_out


def _row_tuple_to_cells(tup: tuple) -> list:
    (
        rank,
        rc,
        rstat,
        uid,
        tg_id,
        tg_un,
        em,
        ev,
        tsc,
        pst,
        sa,
        fa,
        perf,
        wpos,
    ) = tup
    return [
        _csv_cell(rc),
        _csv_cell(rstat),
        rank,
        uid,
        tg_id,
        tg_un or "",
        em or "",
        _csv_cell(ev),
        _csv_cell(tsc),
        _csv_cell(pst),
        _csv_cell(sa),
        _csv_cell(fa),
        "да" if perf else "нет",
        wpos if wpos is not None else "",
    ]


async def export_round_csv(session: AsyncSession, code: RoundCode) -> bytes:
    rnd = await session.scalar(select(Round).where(Round.code == code))
    if rnd is None:
        raise ValueError(f"Round {code.value} not found")

    note = await _export_note(session, rnd)
    tuples = await _build_ordered_score_rows(session, rnd)

    buf = io.StringIO()
    w = csv.writer(buf)
    if note:
        w.writerow([note] + [""] * (len(_SCORE_HEADERS) - 1))
    w.writerow(_SCORE_HEADERS)
    for tup in tuples:
        w.writerow(_row_tuple_to_cells(tup))

    sel = await session.scalar(
        select(WinnerSelection).where(WinnerSelection.round_id == rnd.id)
    )
    if sel is not None:
        w.writerow([])
        w.writerow(["# winners", "selection_id", sel.id])
        wr = await session.execute(
            select(Winner.position, User.telegram_user_id, User.email)
            .join(User, User.id == Winner.user_id)
            .where(Winner.winner_selection_id == sel.id)
            .order_by(Winner.position.asc())
        )
        w.writerow(["position", "telegram_user_id", "email"])
        for pos, tg, em in wr.all():
            w.writerow([pos, tg, em or ""])

    return buf.getvalue().encode("utf-8-sig")


def _row_tuple_to_xlsx_vals(tup: tuple) -> list:
    (
        rank,
        rc,
        rstat,
        uid,
        tg_id,
        tg_un,
        em,
        ev,
        tsc,
        pst,
        sa,
        fa,
        perf,
        wpos,
    ) = tup
    return [
        _xlsx_cell(rc),
        _xlsx_cell(rstat),
        rank,
        uid,
        tg_id,
        tg_un or "",
        em or "",
        _xlsx_cell(ev),
        _xlsx_cell(tsc),
        _xlsx_cell(pst),
        _xlsx_cell(sa),
        _xlsx_cell(fa),
        "да" if perf else "нет",
        wpos if wpos is not None else "",
    ]


async def export_round_xlsx(session: AsyncSession, code: RoundCode) -> bytes:
    rnd = await session.scalar(select(Round).where(Round.code == code))
    if rnd is None:
        raise ValueError(f"Round {code.value} not found")

    note = await _export_note(session, rnd)
    tuples = await _build_ordered_score_rows(session, rnd)

    wb = Workbook()
    sheet = wb.active
    sheet.title = "scores"

    row_cursor = 1
    if note:
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(_SCORE_HEADERS)
        )
        sheet.cell(row=1, column=1, value=note)
        row_cursor = 2

    for col, title in enumerate(_SCORE_HEADERS, start=1):
        sheet.cell(row=row_cursor, column=col, value=title)

    data_start = row_cursor + 1
    for i, tup in enumerate(tuples):
        r = data_start + i
        vals = _row_tuple_to_xlsx_vals(tup)
        for c, v in enumerate(vals, start=1):
            sheet.cell(row=r, column=c, value=v)
        if tup[12]:  # perfect_all
            for c in range(1, len(vals) + 1):
                sheet.cell(row=r, column=c).fill = _FILL_PERFECT

    sel = await session.scalar(
        select(WinnerSelection).where(WinnerSelection.round_id == rnd.id)
    )
    if sel is not None:
        wsh = wb.create_sheet("winners")
        wsh.append(["position", "telegram_user_id", "email"])
        wr = await session.execute(
            select(Winner.position, User.telegram_user_id, User.email)
            .join(User, User.id == Winner.user_id)
            .where(Winner.winner_selection_id == sel.id)
            .order_by(Winner.position.asc())
        )
        for pos, tg, em in wr.all():
            wsh.append([pos, tg, em or ""])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

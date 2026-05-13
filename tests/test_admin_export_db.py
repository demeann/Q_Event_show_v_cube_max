"""Выгрузки туров для админов: CSV / XLSX (SQLite)."""

from __future__ import annotations

import csv
import io
from datetime import datetime

import pytest
from openpyxl import load_workbook

from app.db.models import (
    Round,
    RoundCode,
    RoundStatus,
    User,
    UserRoundProgress,
    Winner,
    WinnerSelection,
)
from app.db.models.progress import RoundProgressStatus
from app.services.admin_export import export_round_csv, export_round_xlsx


@pytest.mark.asyncio
async def test_export_csv_round_missing(db_session):
    with pytest.raises(ValueError, match="R1 not found"):
        await export_round_csv(db_session, RoundCode.R1)


@pytest.mark.asyncio
async def test_export_xlsx_round_missing(db_session):
    with pytest.raises(ValueError, match="R3 not found"):
        await export_round_xlsx(db_session, RoundCode.R3)


async def _seed_round_r1_with_users(db_session, *, with_winners: bool) -> None:
    db_session.add(
        Round(
            id=10,
            code=RoundCode.R1,
            name="T1",
            starts_at=datetime(2026, 6, 1, 0, 0, 0),
            ends_at=datetime(2026, 6, 3, 0, 0, 0),
            status=RoundStatus.ACTIVE,
        )
    )
    db_session.add(
        User(
            id=1,
            telegram_user_id=101,
            email="a@pmru.com",
            email_domain="pmru.com",
            email_verified_at=datetime(2026, 1, 1),
            is_admin=False,
            is_blocked=False,
        )
    )
    db_session.add(
        User(
            id=2,
            telegram_user_id=202,
            email="b@pmru.com",
            email_domain="pmru.com",
            email_verified_at=datetime(2026, 1, 2),
            is_admin=False,
            is_blocked=False,
        )
    )
    db_session.add(
        User(
            id=3,
            telegram_user_id=303,
            email=None,
            email_domain=None,
            email_verified_at=None,
            is_admin=False,
            is_blocked=False,
        )
    )
    db_session.add(
        User(
            id=4,
            telegram_user_id=404,
            email="vip@pmru.com",
            email_domain="pmru.com",
            email_verified_at=None,
            is_admin=True,
            is_blocked=False,
        )
    )
    db_session.add(
        User(
            id=5,
            telegram_user_id=505,
            email="blk@pmru.com",
            email_domain="pmru.com",
            email_verified_at=datetime(2026, 1, 1),
            is_admin=False,
            is_blocked=True,
        )
    )
    await db_session.flush()
    db_session.add(
        UserRoundProgress(
            user_id=1,
            round_id=10,
            status=RoundProgressStatus.FINISHED,
            total_score=42,
        )
    )
    db_session.add(
        UserRoundProgress(
            user_id=2,
            round_id=10,
            status=RoundProgressStatus.IN_PROGRESS,
            total_score=10,
        )
    )
    await db_session.flush()
    if with_winners:
        db_session.add(
            WinnerSelection(
                id=7,
                round_id=10,
                candidates_count=2,
                winners_count=1,
                n_step=0,
                score_threshold=0,
                executed_at=datetime(2026, 6, 4, 12, 0, 0),
                payload={"steps": [], "ordering_strategy": "fixed_list_ranks_v1"},
            )
        )
        await db_session.flush()
        db_session.add(Winner(winner_selection_id=7, user_id=1, position=1))
    await db_session.flush()


@pytest.mark.asyncio
async def test_export_csv_filters_eligible_and_orders_by_score_desc(db_session):
    await _seed_round_r1_with_users(db_session, with_winners=False)
    raw = await export_round_csv(db_session, RoundCode.R1)
    assert raw.startswith(b"\xef\xbb\xbf")
    text = raw.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    header = rows[0]
    idx = {name: i for i, name in enumerate(header)}
    assert idx["total_score"] >= 0
    assert idx["competition_rank"] >= 0
    assert idx["all_answers_correct"] >= 0
    data = [r for r in rows[1:] if r and not r[0].startswith("#")]
    # В рейтинг попадают только участники с записью прогресса по туру.
    assert [r[idx["user_id"]] for r in data] == ["1", "2"]
    assert [r[idx["competition_rank"]] for r in data] == ["1", "2"]
    assert [r[idx["total_score"]] for r in data] == ["42", "10"]
    u1 = next(r for r in data if r[idx["user_id"]] == "1")
    assert u1[idx["telegram_user_id"]] == "101"
    assert u1[idx["email"]] == "a@pmru.com"
    assert u1[idx["progress_status"]] == "FINISHED"
    assert u1[idx["winner_position"]] == ""
    assert u1[idx["round_code"]] == "R1"


@pytest.mark.asyncio
async def test_export_csv_appends_winners_meta_block(db_session):
    await _seed_round_r1_with_users(db_session, with_winners=True)
    raw = await export_round_csv(db_session, RoundCode.R1)
    text = raw.decode("utf-8-sig")
    assert "# winners" in text
    assert "selection_id" in text
    assert "7" in text
    assert "position,telegram_user_id,email" in text.replace("\r\n", "\n")
    assert "101" in text and "a@pmru.com" in text


@pytest.mark.asyncio
async def test_export_xlsx_scores_and_winners_sheets(db_session):
    await _seed_round_r1_with_users(db_session, with_winners=True)
    raw = await export_round_xlsx(db_session, RoundCode.R1)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        assert "scores" in wb.sheetnames
        assert "winners" in wb.sheetnames
        sh = wb["scores"]
        srows = list(sh.iter_rows(values_only=True))
        h = srows[0]
        col_u = h.index("user_id")
        col_sc = h.index("total_score")
        col_rank = h.index("competition_rank")
        data_rows = [r for r in srows[1:] if r and r[col_u] is not None]
        assert [r[col_u] for r in data_rows] == [1, 2]
        assert [r[col_rank] for r in data_rows] == [1, 2]
        assert [r[col_sc] for r in data_rows] == [42, 10]
        assert h[0] == "round_code"
        wh = wb["winners"]
        wrows = list(wh.iter_rows(values_only=True))
        assert wrows[0] == ("position", "telegram_user_id", "email")
        assert wrows[1][0] == 1 and wrows[1][1] == 101
    finally:
        wb.close()
